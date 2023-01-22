# © Зарихин В. А., 2022

import os
import re
from datetime import datetime
from typing import Union

from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from pandas import read_excel

from parser import list_female_names, list_male_names


def get_version():
    return '1.2.4.49'


def get_copyright_sign():
    return '© Зарихин В. А., 2022'


class Parser:
    root_path = os.getcwd()
    pattern_ready_file = 'готовый файл.xlsm'
    pattern_source_folder = 'списки от СК'
    folder_to_move = 'прочитанные файлы'

    pattern_folder_with_names = 'списки имён'
    pattern_female_names_file = 'женские имена.txt'
    pattern_male_names_file = 'мужские имена.txt'

    female_gender = 'Ж'
    male_gender = 'М'

    column_to_write = {
        'Порядковый номер': 1,
        'Фамилия': 2,
        'Имя': 3,
        'Отчество': 4,
        'Пол': 5,
        'Дата рождения': 6,
        'Дата прикрепления': 7,
        'Дата окончания': 8,
        'Дата отмены': 9,
        'Номер полиса': 10,
        'Лимит прикрепления': 11,
        'Наименование договора': 12,
        'Наименование программы': 13,
        'Расширение': 14,
        'Ограничение': 15,
        'Код документа': 16,
        'Серия документа': 17,
        'Номер документа': 18,
        'Кем выдан': 19,
        'Подразделение': 20,
        'Дата выдачи': 21,
        'Телефон пациента': 22,
        'Адрес регистрации': 23,
        'Адрес проживания': 24,
        'СНИЛС': 25,
        'Место работы': 26,
        'Электронная почта': 27
    }

    def __init__(self, folder_to_read: str, dict_to_write: dict = (), sheet_num_to_read=0,
                 start_line_to_read=0, start_column_to_read=0,
                 exclude_column: Union[list, tuple] = (), sep_column: dict = (),
                 step_line=0, extra_cell: dict = (), file_to_write=pattern_ready_file,
                 sheet_num_to_write=0, show_policies=False, show_data=False, save=True,
                 move_after_reading=True):

        self.list_files_to_read = self._get_list_files_to_read(folder_to_read)
        self.sheet_num_to_read = sheet_num_to_read
        self.start_line_to_read = start_line_to_read
        self.start_column_to_read = start_column_to_read

        if exclude_column is None:
            exclude_column = ()
        self.exclude_column = exclude_column

        if sep_column is None:
            sep_column = ()
        self.sep_column = sep_column

        if step_line is None:
            step_line = ()
        self.step_line = step_line

        if extra_cell is None:
            extra_cell = ()
        self.extra_cell = extra_cell

        self.file_to_write = self._validate_file_name(file_to_write)
        self.sheet_num_to_write = sheet_num_to_write
        self.dict_to_write = dict_to_write

        self.female_names_file = self._validate_file_name(self.pattern_female_names_file,
                                                          self.pattern_folder_with_names)
        self.male_names_file = self._validate_file_name(self.pattern_male_names_file,
                                                        self.pattern_folder_with_names)

        self.move_after_reading = move_after_reading
        self.show_data = show_data
        self.show_policies = show_policies
        self.save = save

        self.gender_determined = False

    def _validate_file_name(self, file_name_pattern, folder=None):
        if not folder:
            folder = self.root_path
        try:
            for name in os.listdir(folder):
                if name.lower() == file_name_pattern.lower():
                    return os.path.join(folder, name)
        except FileNotFoundError:
            pass

        return os.path.join(folder, file_name_pattern)

    def _get_list_files_to_read(self, folder_to_read):
        list_files = []

        for source_path in os.listdir(self.root_path):

            if source_path.lower() == self.pattern_source_folder.lower():
                for path_folder_to_read in os.listdir(source_path):

                    if path_folder_to_read.lower() == folder_to_read.lower():
                        path_folder_to_read = os.path.join(source_path, path_folder_to_read)

                        for file_to_read in os.listdir(path_folder_to_read):
                            path_to_file = os.path.join(path_folder_to_read, file_to_read)
                            if os.path.isfile(path_to_file):
                                list_files.append(path_to_file)
                        break

        return list_files

    def get_data_to_write(self):
        list_data = []

        for file_to_read in self.list_files_to_read:
            try:
                data_frame = read_excel(file_to_read, sheet_name=self.sheet_num_to_read)

            except ValueError:
                self.list_files_to_read.remove(file_to_read)
                continue

            except IsADirectoryError:
                self.list_files_to_read.remove(file_to_read)
                continue

            line_to_read = self.start_line_to_read
            last_line = data_frame.shape[0]

            start_column = self.start_column_to_read
            last_column = data_frame.shape[1]

            start_table = line_to_read
            table_start_value = str(data_frame.iloc[self.start_line_to_read - 1, start_column])

            while line_to_read < last_line:
                val_line = str(data_frame.iloc[line_to_read, start_column])
                if val_line == 'nan' or val_line.isspace():
                    if self.step_line:
                        line_to_read += self.step_line
                    else:
                        line_to_read = self._get_next_table(data_frame,
                                                            last_line,
                                                            line_to_read,
                                                            start_column,
                                                            table_start_value)

                    if line_to_read >= last_line - 1:
                        break

                    start_table = line_to_read
                    val_line = str(data_frame.iloc[line_to_read, start_column])
                    if val_line == 'nan' or val_line.isspace():
                        break

                data_line = []
                for num_column in range(start_column, last_column):
                    if num_column in self.exclude_column:
                        continue

                    cell_value = str(data_frame.iloc[line_to_read, num_column])

                    if cell_value == 'nan' or cell_value.isspace():
                        self._append_value_to_data_line(data_line, None)
                        continue

                    try:
                        cell_value = datetime.strptime(cell_value,
                                                       '%Y-%m-%d %H:%M:%S').strftime('%d.%m.%Y')
                    except ValueError:
                        pass

                    if num_column in self.sep_column:
                        sep = self.sep_column.get(num_column)
                        cell_value = self._split_value(cell_value, sep)

                    self._append_value_to_data_line(data_line, self._determine_gender(cell_value))

                for key in self.extra_cell:
                    line, col = key.split()
                    sep = self.extra_cell[key]
                    cell_value = str(data_frame.iloc[start_table - int(line), int(col)])

                    if cell_value and sep:
                        if type(sep) == bool:
                            cell_value = re.split('\n|: |С |По |с |по | Г. | г.', cell_value)
                        else:
                            cell_value = re.split(sep, cell_value)

                    self._append_value_to_data_line(data_line, cell_value)

                if not self.gender_determined:
                    gender = self._get_gender_from_lists_of_names(data_line)
                    self._append_value_to_data_line(data_line, gender)

                list_data.append(data_line)
                line_to_read += 1

            if self.move_after_reading:
                self._move_after_reading(file_to_read)

        return list_data

    def _move_after_reading(self, file_to_move):
        source_path, file = os.path.split(file_to_move)
        folder_to_move = os.path.join(source_path, self.folder_to_move)
        path_to_move = os.path.join(folder_to_move, file)

        self._create_folder(folder_to_move)

        try:
            os.rename(file_to_move, path_to_move)
        except FileExistsError:
            pass

    @staticmethod
    def _create_folder(path_folder):
        try:
            os.makedirs(path_folder)
        except FileExistsError:
            pass

    @staticmethod
    def _get_next_table(data_frame, last_line, line_to_read, column, table_start_value):
        for line in range(line_to_read, last_line):
            val_line = str(data_frame.iloc[line, column])
            if val_line == table_start_value:
                return line + 1

            if line >= last_line - 1:
                return line

    @staticmethod
    def _split_value(cell_value, sep=None):
        exclude_sep = [' ', '\n']

        cell_value = cell_value.split(sep=sep)

        if len(cell_value) == 2 and not sep:
            cell_value.append('')

        if sep and sep not in exclude_sep:
            cell_value[-1] = sep + cell_value[-1]

        return cell_value

    @staticmethod
    def _append_value_to_data_line(data_line: list, values):
        if type(values) == list:
            for val in values:
                data_line.append(val.title().strip())
        else:
            try:
                data_line.append(values.title().strip())
            except AttributeError:
                data_line.append(values)

    def create_file_to_write(self):
        new_file_to_write = Workbook()
        writable_sheet = new_file_to_write.worksheets[0]

        for value, num_column in self.column_to_write.items():
            writable_sheet.cell(row=1, column=num_column).value = value

        new_file_to_write.save(self.file_to_write)

        print(f'Created file to write "{self.file_to_write}"')

    def write_data(self, data_to_write):
        try:
            writable_file = load_workbook(self.file_to_write, read_only=False, keep_vba=True)
        except FileNotFoundError:
            self.create_file_to_write()
            writable_file = load_workbook(self.file_to_write, read_only=False, keep_vba=True)

        writable_sheet = writable_file.worksheets[self.sheet_num_to_write]

        last_line_the_file = writable_sheet.max_row
        line_to_write = last_line_the_file + 1
        first_column = writable_sheet.min_column

        policies = self.get_list_policies(writable_sheet=writable_sheet)
        if self.show_policies:
            print(f'policies => {policies}')

        policy_position = self.dict_to_write['Номер полиса']
        for idx_line, line in enumerate(data_to_write):
            if line[policy_position] in policies:
                if self.show_policies:
                    print(f'The policy: "{line[policy_position]}" is already in the recorded file')
                continue

            writable_sheet.cell(row=line_to_write + idx_line,
                                column=first_column).value = last_line_the_file + idx_line

            for key, num_value in self.dict_to_write.items():
                column_to_write = self.column_to_write.get(key)
                value = None

                try:
                    value = line[num_value]
                except IndexError:
                    pass

                writable_sheet.cell(row=line_to_write + idx_line,
                                    column=column_to_write).value = value

        if self.save:
            self._save_file_to_exel(writable_file)
        else:
            print(f'SAVE = {self.save}')

    @staticmethod
    def print_data_for_line(data):
        for i, line in enumerate(data):
            print()
            print(f'Data line "{i}"')

            for val in enumerate(line):
                print(val)

    def _determine_gender(self, val):
        female_gender_list = ['WOMEN', 'ЖЕНСКИЙ', 'ЖЕН',
                              'Women', 'Женский', 'Жен', 'Ж',
                              'women', 'женский', 'жен', 'ж']

        male_gender_list = ['MEN', 'МУЖСКОЙ', 'МУЖ',
                            'Men', 'Мужской', 'Муж', 'М',
                            'men', 'мужской', 'муж', 'м']

        if val in female_gender_list:
            self.gender_determined = True
            return self.female_gender

        if val in male_gender_list:
            self.gender_determined = True
            return self.male_gender

        return val

    def _get_gender_from_lists_of_names(self, data_line):
        try:
            with open(self.female_names_file, 'r', encoding='utf-8') as female:
                female_names = female.read()
        except FileNotFoundError:
            female_names = ()

        try:
            with open(self.male_names_file, 'r', encoding='utf-8') as male:
                male_names = male.read()
        except FileNotFoundError:
            male_names = ()

        val = data_line[self.dict_to_write['Имя']]
        if not val:
            return

        if val in list_female_names or val in female_names:
            return self.female_gender

        if val in list_male_names or val in male_names:
            return self.male_gender

    def get_list_policies(self, writable_sheet):
        start_line_file = 2
        last_line_file = writable_sheet.max_row
        policies_column = self.column_to_write['Номер полиса']

        list_policies = []
        for line_pol in range(start_line_file, last_line_file + 1):
            value_cell = writable_sheet.cell(row=line_pol, column=policies_column).value
            list_policies.append(str(value_cell))

        return list_policies

    def _save_file_to_exel(self, writable_file):
        writable_file.save(self.file_to_write)
        name_file_to_write = os.path.basename(self.file_to_write)

        file_names_to_read = []
        for path_file in self.list_files_to_read:
            file_names_to_read.append(os.path.basename(path_file))

        print(f'Data from "{file_names_to_read}" is written to "{name_file_to_write}"!')

    def pars(self):
        try:
            if self.list_files_to_read:
                data_to_write = self.get_data_to_write()

                if self.show_data:
                    self.print_data_for_line(data_to_write)

                self.write_data(data_to_write=data_to_write)

        except FileNotFoundError as file_not_found:
            print(f'File not found! {file_not_found}')

        except KeyError as key_error:
            print(f'Key "{key_error}" not found!')

        except TypeError as type_error:
            print(f'type_error! {type_error}')

    def copy_to_csv_format(self, source_file: str, path_to_save='./csv_files', sheet_num=0):
        self._create_folder(path_to_save)

        path_source_file = None
        for file in self.list_files_to_read:
            if file.split(sep='/')[-1] == source_file:
                path_source_file = file

        if path_source_file:
            to_csv_file = f'{path_to_save}/{source_file}.csv'
            read_excel(path_source_file, sheet_name=sheet_num).to_csv(to_csv_file)
            print(f'"{source_file}" copied to CSV format!')
        else:
            print(f'File "{source_file}" not found!')
