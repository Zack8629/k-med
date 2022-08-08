import os
from datetime import datetime
from typing import Union

from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from pandas import read_excel


class Parser:
    root_path = os.getcwd()
    pattern_ready_file = 'готовый файл.xlsm'
    pattern_source_folder = 'списки от СК'

    pattern_folder_with_names = 'списки имён'
    pattern_female_names_file = 'женские имена.txt'
    pattern_male_names_file = 'мужские имена.txt'

    female_gender = 'Ж'
    male_gender = 'М'

    column_to_write = {
        'Порядковый номер': 1,
        'Фамилия': 2,
        'Имя': 3,
        'Отчество': 4,
        'Пол': 5,
        'Дата рождения': 6,
        'Дата прикрепления': 7,
        'Дата окончания': 8,
        'Дата отмены': 9,
        'Номер полиса': 10,
        'Лимит прикрепления': 11,
        'Наименование договора': 12,
        'Наименование программы': 13,
        'Расширение': 14,
        'Ограничение': 15,
        'Код документа': 16,
        'Серия документа': 17,
        'Номер документа': 18,
        'Кем выдан': 19,
        'Подразделение': 20,
        'Дата выдачи': 21,
        'Телефон пациента': 22,
        'Адрес регистрации': 23,
        'Адрес проживания': 24,
        'СНИЛС': 25,
        'Место работы': 26,
        'Электронная почта': 27
    }

    def __init__(self, folder_to_read: str, dict_to_write: dict = (), sheet_num_to_read=0,
                 start_line_to_read=0, start_column_to_read=0,
                 exclude_column: Union[list, tuple] = (), sep_column: dict = (),
                 step_line=0, extra_cell: dict = (), file_to_write=pattern_ready_file,
                 sheet_num_to_write=0, show_policies=False, show_data=False, save=True):

        self.list_files_to_read = self._get_list_files_to_read(folder_to_read)
        self.sheet_num_to_read = sheet_num_to_read
        self.start_line_to_read = start_line_to_read
        self.start_column_to_read = start_column_to_read
        self.exclude_column = exclude_column
        self.sep_column = sep_column
        self.step_line = step_line
        self.extra_cell = extra_cell

        self.file_to_write = self._validate_file_name(file_to_write)
        self.sheet_num_to_write = sheet_num_to_write
        self.dict_to_write = dict_to_write

        self.female_names_file = self._validate_file_name(self.pattern_female_names_file,
                                                          self.pattern_folder_with_names)
        self.male_names_file = self._validate_file_name(self.pattern_male_names_file,
                                                        self.pattern_folder_with_names)

        self.show_data = show_data
        self.show_policies = show_policies
        self.save = save

        self.gender_determined = False

    def _validate_file_name(self, file_name_pattern, folder=None):
        if not folder:
            folder = self.root_path
        try:
            for name in os.listdir(folder):
                if name.lower() == file_name_pattern.lower():
                    return os.path.join(folder, name)
        except FileNotFoundError:
            pass

        return os.path.join(folder, file_name_pattern)

    def _get_list_files_to_read(self, folder_to_read):
        list_files = []

        for source_path in os.listdir(self.root_path):

            if source_path.lower() == self.pattern_source_folder.lower():
                for path_folder_to_read in os.listdir(source_path):

                    if path_folder_to_read.lower() == folder_to_read.lower():
                        path_folder_to_read = os.path.join(source_path, path_folder_to_read)

                        for file_to_read in os.listdir(path_folder_to_read):
                            path_to_file = os.path.join(path_folder_to_read, file_to_read)
                            list_files.append(path_to_file)
                        break

        return list_files

    def get_data_to_write(self):
        list_data = []

        for file_to_read in self.list_files_to_read:
            try:
                data_frame = read_excel(file_to_read, sheet_name=self.sheet_num_to_read)
            except ValueError:
                self.list_files_to_read.remove(file_to_read)
                continue

            line_to_read = self.start_line_to_read
            last_line = data_frame.shape[0]

            start_column = self.start_column_to_read
            last_column = data_frame.shape[1]

            while line_to_read < last_line:
                val_line = str(data_frame.iloc[line_to_read, start_column])
                if val_line == 'nan' or val_line.isspace():
                    line_to_read += self.step_line
                    if line_to_read > last_line:
                        break

                    val_line = str(data_frame.iloc[line_to_read, start_column])
                    if val_line == 'nan' or val_line.isspace():
                        break

                data_line = []
                for num_column in range(start_column, last_column):
                    if num_column in self.exclude_column:
                        continue

                    cell_value = str(data_frame.iloc[line_to_read, num_column])

                    if cell_value == 'nan' or cell_value.isspace():
                        self._append_value_to_data_line(data_line, None)
                        continue

                    try:
                        cell_value = datetime.strptime(cell_value,
                                                       '%Y-%m-%d %H:%M:%S').strftime('%d.%m.%Y')
                    except ValueError:
                        pass

                    if num_column in self.sep_column:
                        sep = self.sep_column.get(num_column)
                        cell_value = cell_value.split(sep=sep)

                        if len(cell_value) == 2 and not sep:
                            cell_value.append('')

                        if sep and sep != ' ':
                            cell_value[-1] = sep + cell_value[-1]

                    self._append_value_to_data_line(data_line, self._determine_gender(cell_value))

                for key in self.extra_cell:
                    line, col = key.split()
                    cell_value = str(data_frame.iloc[int(line), int(col)])

                    if self.extra_cell[key]:
                        cell_value = cell_value.split()

                    self._append_value_to_data_line(data_line, cell_value)

                if not self.gender_determined:
                    gender = self._get_gender_from_lists_of_names(data_line)
                    self._append_value_to_data_line(data_line, gender)

                list_data.append(data_line)
                line_to_read += 1

        return list_data

    @staticmethod
    def _append_value_to_data_line(data_line: list, values):
        if type(values) == list:
            for val in values:
                data_line.append(val.title())
        else:
            try:
                data_line.append(values.title())
            except AttributeError:
                data_line.append(values)

    def create_file_to_write(self):
        new_file_to_write = Workbook()
        writable_sheet = new_file_to_write.worksheets[0]

        for value, num_column in self.column_to_write.items():
            writable_sheet.cell(row=1, column=num_column).value = value

        new_file_to_write.save(self.file_to_write)

        print(f'Created file to write "{self.file_to_write}"')

    def write_data(self, data_to_write):
        try:
            writable_file = load_workbook(self.file_to_write, read_only=False, keep_vba=True)
        except FileNotFoundError:
            self.create_file_to_write()
            writable_file = load_workbook(self.file_to_write, read_only=False, keep_vba=True)

        writable_sheet = writable_file.worksheets[self.sheet_num_to_write]

        last_line_the_file = writable_sheet.max_row
        line_to_write = last_line_the_file + 1
        first_column = writable_sheet.min_column

        policies = self.get_list_policies(writable_sheet=writable_sheet)
        if self.show_policies:
            print(f'policies => {policies}')

        policy_position = self.dict_to_write['Номер полиса']
        for idx_line, line in enumerate(data_to_write):
            if line[policy_position] in policies:
                if self.show_policies:
                    print(f'The policy: "{line[policy_position]}" is already in the recorded file')
                continue

            writable_sheet.cell(row=line_to_write + idx_line,
                                column=first_column).value = last_line_the_file + idx_line

            for key, num_value in self.dict_to_write.items():
                column_to_write = self.column_to_write.get(key)
                value = None

                try:
                    value = line[num_value]
                except IndexError:
                    pass

                writable_sheet.cell(row=line_to_write + idx_line,
                                    column=column_to_write).value = value

        if self.save:
            self._save_file_to_exel(writable_file)
        else:
            print(f'SAVE = {self.save}')

    @staticmethod
    def print_data_for_line(data):
        for i, line in enumerate(data):
            print()
            print(f'Data line "{i}"')

            for val in enumerate(line):
                print(val)

    def _determine_gender(self, val):
        female_gender_list = ['WOMEN', 'ЖЕНСКИЙ', 'ЖЕН',
                              'Women', 'Женский', 'Жен', 'Ж',
                              'women', 'женский', 'жен', 'ж']

        male_gender_list = ['MEN', 'МУЖСКОЙ', 'МУЖ',
                            'Men', 'Мужской', 'Муж', 'М',
                            'men', 'мужской', 'муж', 'м']

        if val in female_gender_list:
            self.gender_determined = True
            return self.female_gender

        if val in male_gender_list:
            self.gender_determined = True
            return self.male_gender

        return val

    def _get_gender_from_lists_of_names(self, data_line):
        try:
            with open(self.female_names_file, 'r', encoding='utf-8') as female:
                female_names = female.read()
        except FileNotFoundError:
            female_names = ()

        try:
            with open(self.male_names_file, 'r', encoding='utf-8') as male:
                male_names = male.read()
        except FileNotFoundError:
            male_names = ()

        list_female_names = ['Ава', 'Августа', 'Аврелия', 'Аврора', 'Агата', 'Агафья', 'Агнес',
                             'Агнесса', 'Агния', 'Аделаида', 'Аделина', 'Адриенна', 'Азиза',
                             'Аида', 'Айгуль', 'Алдона', 'Алевтина', 'Александра', 'Алима',
                             'Алина', 'Алиса', 'Алия', 'Алла', 'Альбина', 'Аманда', 'Амина',
                             'Амира', 'Анаис', 'Анастасия', 'Ангелина', 'Анжела', 'Анжелика',
                             'Анисия', 'Анна', 'Антонина', 'Анук', 'Анфиса', 'Аполлинария',
                             'Аполлония', 'Арабелла', 'Ариана', 'Ассоль', 'Аурелия', 'Бажена',
                             'Беата', 'Беатриса', 'Белла', 'Блерта', 'Божена', 'Валентина',
                             'Валерия', 'Ванда', 'Варвара', 'Василиса', 'Венди', 'Вера',
                             'Вероника', 'Виктория', 'Виолетта', 'Галатея', 'Галина', 'Глафира',
                             'Гликерия', 'Гоар', 'Говхар', 'Горица', 'Гульмира', 'Гульнара',
                             'Гуннхильда', 'Гюльджан', 'Дана', 'Дарерка', 'Дарья', 'Дебора',
                             'Джанет', 'Дженифер', 'Дженна', 'Джессика', 'Джоан', 'Джулия',
                             'Диана', 'Дина', 'Дита', 'Домна', 'Дора', 'Доротея', 'Ева', 'Евгения',
                             'Евдокия', 'Евдоксия', 'Евлалия', 'Евлампия', 'Евпраксия',
                             'Екатерина', 'Елена', 'Елизавета', 'Епистима', 'Ермиония', 'Жасмин',
                             'Женевьева', 'Забава', 'Зинаида', 'Зоя', 'Зульфия', 'Ивета', 'Илона',
                             'Ильзе', 'Инга', 'Индира', 'Инес', 'Инна', 'Иоанна', 'Ираида',
                             'Ирина', 'Ирма', 'Иулия', 'Ия', 'Йенни', 'Камилла', 'Камиля', 'Карен',
                             'Карина', 'Каролина', 'Ким', 'Кира', 'Кирсти', 'Клавдия', 'Клара',
                             'Клементина', 'Констанция', 'Консуэло', 'Корнелия', 'Кристина',
                             'Ксения', 'Лада', 'Лана', 'Лаодика', 'Лариса', 'Лаура', 'Лейла',
                             'Леля', 'Лидия', 'Лина', 'Линнея', 'Лия', 'Лора', 'Лукия', 'Любовь',
                             'Людмила', 'Людовика', 'Магали', 'Магдалина', 'Мадина', 'Майя',
                             'Малика', 'Мальвина', 'Маргарет', 'Маргарита', 'Марианна', 'Марина',
                             'Мариса', 'Марисоль', 'Мария', 'Марлен', 'Марфа', 'Мастридия',
                             'Матильда', 'Матрёна', 'Мафтуха', 'Мелания', 'Мелисса', 'Меропа',
                             'Мерседес', 'Милица', 'Миранда', 'Мирей', 'Миропия', 'Мирослава',
                             'Михримах', 'Мэдисон', 'Мэри', 'Мю', 'Надежда', 'Наджия', 'Надия',
                             'Назгуль', 'Наиля', 'Наоми', 'Наталья', 'Невена', 'Нелли', 'Неонилла',
                             'Ника', 'Николь', 'Николетта', 'Нилуфар', 'Нинель', 'Ноа', 'Нонна',
                             'Нора', 'Нэнси', 'Одетта', 'Октябрина', 'Олимпиада', 'Ольга',
                             'Ориана', 'Павла', 'Павлина', 'Параскева', 'Пинна', 'Пнина', 'Полина',
                             'Прасковья', 'Прити', 'Рада', 'Раиса', 'Рамина', 'Раминта', 'Рамона',
                             'Рамона', 'Ребекка', 'Ревекка', 'Регина', 'Римма', 'Рогнеда', 'Роза',
                             'Розалия', 'Рос', 'Росарио', 'Рукайя', 'Руслана', 'Руфина', 'Рушан',
                             'Сабина', 'Саида', 'Салиха', 'Саманта', 'Сандра', 'Сара', 'Светлана',
                             'Серафима', 'Сибилла', 'Сильвия', 'Синклитикия', 'Синтия', 'Смиляна',
                             'Снежана', 'Сона-Ханум', 'Соня', 'София', 'Стелла', 'Степанида',
                             'Стефания', 'Тавус', 'Тамара', 'Танзиля', 'Тарья', 'Татьяна',
                             'Тахмина', 'Томоми', 'Ульяна', 'Урсула', 'Урсула', 'Фаина',
                             'Фарангис', 'Фатима', 'Феба', 'Фейт', 'Фёкла', 'Фелисити', 'Феодосия',
                             'Феофания', 'Фива', 'Фила', 'Филлида', 'Фотина', 'Франческа', 'Фрида',
                             'Ханнелора', 'Хатидже', 'Хафса', 'Хильдегарда', 'Хильдур', 'Цветана',
                             'Целестина', 'Цецилия', 'Чулпан', 'Шарлотта', 'Шейла', 'Шерил',
                             'Шорена', 'Эдита', 'Эдна', 'Элеонора', 'Элла', 'Эллен', 'Эльвира',
                             'Эльмира', 'Эми', 'Эмилия', 'Эмма', 'Эннафа', 'Эрвина', 'Эрика',
                             'Эрин', 'Эрна', 'Этель', 'Юлия', 'Юния', 'Яна', 'Ярослава',
                             'Арина', 'Оксана', 'Наталия', 'Алена', 'Дарина', 'Наталия', 'Альмира']

        list_male_names = ['Аарон', 'Аббас', 'Абд', 'Абдуллах', 'Абид', 'Аботур', 'Аввакум',
                           'Август', 'Авдей', 'Авель', 'Авигдор', 'Авксентий', 'Авл', 'Авнер',
                           'Аврелий', 'Автандил', 'Автоном', 'Агапит', 'Агафангел', 'Агафодор',
                           'Агафон', 'Агриппа', 'Адам', 'Адам', 'Адар', 'Адиль', 'Адольф',
                           'Адонирам', 'Адриан', 'Азамат', 'Азат', 'Азиз', 'Азим', 'Айварс',
                           'Айдар', 'Акакий', 'Аквилий', 'Акиф', 'Акоп', 'Аксель', 'Алан', 'Алан',
                           'Аланус', 'Александр', 'Алексей', 'Алик', 'Алим', 'Алипий', 'Алишер',
                           'Алоиз', 'Альберик', 'Альберт', 'Альбин', 'Альваро', 'Альвиан',
                           'Альвизе', 'Альфонс', 'Альфред', 'Амадис', 'Амвросий', 'Амедей', 'Амин',
                           'Амир', 'Амр', 'Анания', 'Анас', 'Анастасий', 'Анатолий', 'Андокид',
                           'Андрей', 'Андроник', 'Аникита', 'Аннерс', 'Анри', 'Ансельм', 'Антипа',
                           'Антон', 'Антоний', 'Антонин', 'Арам', 'Арефа', 'Арзуман', 'Аристарх',
                           'Ариф', 'Аркадий', 'Арсен', 'Арсений', 'Артём', 'Артемий', 'Артур',
                           'Арфаксад', 'Архипп', 'Атанасий', 'Аттик', 'Афанасий', 'Афинагор',
                           'Афиней', 'Африкан', 'Ахилл', 'Ахмад', 'Ахтям', 'Ашот', 'Бадр', 'Барни',
                           'Бартоломео', 'Басир', 'Бахтияр', 'Бен', 'Бехруз', 'Билял', 'Богдан',
                           'Болеслав', 'Болот', 'Бонавентура', 'Борис', 'Борислав', 'Боян',
                           'Бронислав', 'Брячислав', 'Булат', 'Бурхан', 'Бямбасурэн', 'Вадим',
                           'Валентин', 'Валерий', 'Валерьян', 'Вальдемар', 'Вангьял', 'Варлам',
                           'Варнава', 'Варсонофий', 'Варфоломей', 'Василий', 'Вахтанг', 'Велвел',
                           'Велимир', 'Венансио', 'Вениамин', 'Венцеслав', 'Верослав', 'Викентий',
                           'Виктор', 'Викторин', 'Вильгельм', 'Винцас', 'Виссарион', 'Виталий',
                           'Витаутас', 'Вито', 'Владимир', 'Владислав', 'Владлен', 'Влас', 'Волк',
                           'Володарь', 'Вольфганг', 'Вописк', 'Всеволод', 'Всеслав', 'Вук',
                           'Вукол', 'Вышеслав', 'Вячеслав', 'Габриеле', 'Гавриил', 'Гай',
                           'Галактион', 'Гарет', 'Гаспар', 'Гафур', 'Гвидо', 'Гейдар', 'Геласий',
                           'Гельмут', 'Геннадий', 'Генри', 'Генрих', 'Георге', 'Георгий',
                           'Гераклид', 'Герберт', 'Герман', 'Германн', 'Геронтий', 'Герхард',
                           'Гессий', 'Гильем', 'Гинкмар', 'Глеб', 'Гней', 'Гонорий', 'Горацио',
                           'Гордей', 'Гостомысл', 'Градислав', 'Григорий', 'Гримоальд', 'Груди',
                           'Гуго', 'Гьялцен', 'Давид', 'Далер', 'Дамдинсурэн', 'Дамир', 'Данакт',
                           'Даниил', 'Дарий', 'Демид', 'Демьян', 'Денис', 'Децим', 'Джаббар',
                           'Джамиль', 'Джанер', 'Джанфранко', 'Джаррах', 'Джафар', 'Джейкоб',
                           'Джихангир', 'Джованни', 'Джон', 'Джулиус', 'Диодор', 'Диомид',
                           'Дмитрий', 'Доминик', 'Дональд', 'Донат', 'Дорофей', 'Досифей',
                           'Еварест', 'Евгений', 'Евграф', 'Евдоким', 'Евпатий', 'Евсей',
                           'Евстафий', 'Евтихиан', 'Евтихий', 'Евферий', 'Егор', 'Елеазар',
                           'Елисей', 'Емельян', 'Ерванд', 'Еремей', 'Ермак', 'Ермолай', 'Ернар',
                           'Ерофей', 'Ефим', 'Ефрем', 'Жан', 'Ждан', 'Жером', 'Жоан', 'Жюль',
                           'Завид', 'Зайнуддин', 'Закир', 'Захар', 'Захария', 'Збигнев', 'Зденек',
                           'Зеэв', 'Зигмунд', 'Зия', 'Золтан', 'Зосима', 'Зураб', 'Иан', 'Ибрахим',
                           'Иван', 'Иван', 'Ивар', 'Игнатий', 'Игорь', 'Иероним', 'Иерофей',
                           'Изот', 'Израиль', 'Икрима', 'Иларий', 'Илларион', 'Илья', 'Иоаким',
                           'Иоанн', 'Иоанникий', 'Иоахим', 'Иов', 'Иоганнес', 'Ионафан', 'Иосафат',
                           'Ираклий', 'Иржи', 'Иринарх', 'Ириней', 'Иродион', 'Иса', 'Иса',
                           'Исаак', 'Исаия', 'Ислам', 'Исмаил', 'Истислав', 'Истома', 'Истукарий',
                           'Иуда', 'Иулиан', 'Иштван', 'Кадваллон', 'Кадир', 'Казимир', 'Каликст',
                           'Калин', 'Каллистрат', 'Кальман', 'Камран', 'Капитон', 'Карен',
                           'Картерий', 'Касим', 'Кассиан', 'Кассий', 'Касторий', 'Квинт', 'Кехлер',
                           'Киллиан', 'Ким', 'Кир', 'Кирилл', 'Клаас', 'Клавдиан', 'Клеоник',
                           'Климент', 'Кондрат', 'Конон', 'Конрад', 'Константин', 'Корнелиус',
                           'Корнилий', 'Коррадо', 'Косьма', 'Кратипп', 'Криспин', 'Кристиан',
                           'Кронид', 'Кузьма', 'Куприян', 'Курбан', 'Курт', 'Кутлуг-Буга', 'Кэлин',
                           'Лаврентий', 'Лаврентий', 'Лавс', 'Ладислав', 'Лазарь', 'Лайл',
                           'Лампрехт', 'Ландульф', 'Лев', 'Левенте', 'Леви', 'Ленни', 'Леонид',
                           'Леонтий', 'Леонхард', 'Лиам', 'Линкей', 'Логгин', 'Лоренц', 'Лоренцо',
                           'Луи', 'Луитпольд', 'Лука', 'Лукий', 'Лукьян', 'Луций', 'Людовик',
                           'Люцифер', 'Майнхард', 'Макар', 'Макарий', 'Максим', 'Максимиан',
                           'Максимилиан', 'Малик', 'Малх', 'Мамбет', 'Мамонт', 'Маний', 'Маноле',
                           'Мануил', 'Мануэль', 'Мариан', 'Марк', 'Маркел', 'Мартын', 'Марчелло',
                           'Матвей', 'Матео', 'Матиас', 'Матфей', 'Матфий', 'Махмуд', 'Меир',
                           'Мелентий', 'Мелитон', 'Менахем', 'Месроп', 'Мефодий', 'Мечислав',
                           'Мика', 'Микола', 'Микулаш', 'Милорад', 'Милутин', 'Мина', 'Мирко',
                           'Митрофан', 'Михаил', 'Младан', 'Модест', 'Моисей', 'Мордехай',
                           'Мстислав', 'Мурад', 'Мухаммед', 'Мэдисон', 'Мэлс', 'Назар', 'Наиль',
                           'Насиф', 'Натан', 'Натаниэль', 'Наум', 'Нафанаил', 'Нацагдорж',
                           'Нестор', 'Никандр', 'Никанор', 'Никита', 'Никифор', 'Никодим',
                           'Николай', 'Нил', 'Нильс', 'Ноа', 'Ной', 'Норд', 'Оге', 'Одинец',
                           'Октавий', 'Олаф', 'Оле', 'Олег', 'Оливер', 'Оливер', 'Ольгерд',
                           'Онисим', 'Онуфрий', 'Орал', 'Орест', 'Осип', 'Оскар', 'Осман', 'Оттон',
                           'Очирбат', 'Пабло', 'Павел', 'Павлин', 'Павсикакий', 'Паисий',
                           'Палладий', 'Панкратий', 'Пантелеймон', 'Папа', 'Паруйр', 'Парфений',
                           'Патрик', 'Пафнутий', 'Пахомий', 'Педро', 'Перец', 'Пётр', 'Пимен',
                           'Пинхас', 'Пипин', 'Питирим', 'Платон', 'Полидор', 'Полиевкт',
                           'Поликарп', 'Поликрат', 'Порфирий', 'Потап', 'Предраг', 'Премысл',
                           'Пров', 'Прокл', 'Прокул', 'Протасий', 'Прохор', 'Публий', 'Рагнар',
                           'Рагуил', 'Радмир', 'Радослав', 'Раймонд', 'Рамадан', 'Рамазан',
                           'Рамиль', 'Ратмир', 'Рахман', 'Рашад', 'Рашид', 'Рейнгард', 'Рейнхард',
                           'Ренат', 'Реститут', 'Ричард', 'Роберт', 'Родерик', 'Родион',
                           'Родослав', 'Рожер', 'Розарио', 'Роман', 'Рон', 'Ронан', 'Ростислав',
                           'Рудольф', 'Руслан', 'Руф', 'Руфин', 'Рушан', 'Рюрик', 'Сабит',
                           'Сабриэль', 'Савва', 'Савватий', 'Савелий', 'Савин', 'Саддам', 'Садик',
                           'Саид', 'Салават', 'Салих', 'Саллюстий', 'Самад', 'Самуил', 'Сармат',
                           'Сасоний', 'Святослав', 'Северин', 'Секст', 'Семён', 'Септимий',
                           'Серапион', 'Сергей', 'Серж', 'Сигеберт', 'Сила', 'Сильвестр', 'Симеон',
                           'Симон', 'Созон', 'Соломон', 'Сонам', 'Софрон', 'Спиридон', 'Срджан',
                           'Станислав', 'Степан', 'Стефано', 'Стивен', 'Сулейман', 'Сфенел',
                           'Таврион', 'Тавус', 'Тагир', 'Тадеуш', 'Тарас', 'Тарасий', 'Теймураз',
                           'Тейс', 'Тендзин', 'Терентий', 'Терри', 'Тиберий', 'Тигран', 'Тимофей',
                           'Тимур', 'Тимур', 'Тихомир', 'Тихон', 'Томоми', 'Торос', 'Тофик',
                           'Трифон', 'Тудхалия', 'Тур', 'Тутмос', 'Тьерри', 'Уве', 'Уильям',
                           'Улдис', 'Ульф', 'Урбан', 'Урызмаг', 'Усама', 'Усман', 'Фавст',
                           'Фаддей', 'Фадлалла', 'Фарид', 'Фахраддин', 'Федериго', 'Фёдор',
                           'Федосей', 'Федот', 'Фейсал', 'Феликс', 'Феоктист', 'Феофан', 'Феофил',
                           'Феофилакт', 'Фердинанд', 'Ференц', 'Филарет', 'Филип', 'Филипп',
                           'Философ', 'Филострат', 'Фока', 'Фома', 'Фотий', 'Франц', 'Франческо',
                           'Фредерик', 'Фридрих', 'Фродо', 'Фрол', 'Фульк', 'Хайме', 'Ханс',
                           'Харальд', 'Харитон', 'Хасан', 'Хетаг', 'Хильдерик', 'Хирам', 'Хлодвиг',
                           'Хокон', 'Хоселито', 'Хосрой', 'Хотимир', 'Хрисанф', 'Христофор',
                           'Цэрэндорж', 'Чеслав', 'Шалом', 'Шамиль', 'Шамсуддин', 'Шапур', 'Шарль',
                           'Шейх-Хайдар', 'Шон', 'Эберхард', 'Эвандр', 'Эдмунд', 'Эдна', 'Эдуард',
                           'Элбэгдорж', 'Элджернон', 'Элиас', 'Эллиот', 'Эмиль', 'Энвер', 'Энрик',
                           'Энрико', 'Энтони', 'Эразм', 'Эрик', 'Эрик', 'Эрнст', 'Эстебан',
                           'Этьен', 'Ювеналий', 'Юлиан', 'Юлий', 'Юлиус', 'Юрген', 'Юрий', 'Юстин',
                           'Юхан', 'Яков', 'Якуб', 'Ян', 'Яни', 'Януарий', 'Яромир', 'Ярополк',
                           'Ярослав',
                           'Артем', 'Ахмед', 'Данил', 'Кеннет', 'Марат', 'Петр', 'Рустем',
                           'Салихат', 'Федор']

        val = data_line[self.dict_to_write['Имя']]
        if not val:
            return

        if val in list_female_names or val in female_names:
            return self.female_gender

        if val in list_male_names or val in male_names:
            return self.male_gender

    def get_list_policies(self, writable_sheet):
        start_line_file = 2
        last_line_file = writable_sheet.max_row
        policies_column = self.column_to_write['Номер полиса']

        list_policies = []
        for line_pol in range(start_line_file, last_line_file + 1):
            value_cell = writable_sheet.cell(row=line_pol, column=policies_column).value
            list_policies.append(str(value_cell))

        return list_policies

    def _save_file_to_exel(self, writable_file):
        writable_file.save(self.file_to_write)
        name_file_to_write = self.file_to_write.split('/')[-1]

        file_names_to_read = []
        for path_file in self.list_files_to_read:
            file_names_to_read.append(path_file.split('/')[-1])

        print(f'Data from "{file_names_to_read}" is written to "{name_file_to_write}"!')

    def pars(self):
        try:
            data_to_write = self.get_data_to_write()

            if self.show_data:
                self.print_data_for_line(data_to_write)

            self.write_data(data_to_write=data_to_write)

        except FileNotFoundError as file_not_found:
            print(f'File not found! {file_not_found}')

        except KeyError as key_error:
            print(f'Key "{key_error}" not found!')

        except TypeError as type_error:
            print(f'type_error! {type_error}')

    def copy_to_csv_format(self, source_file: str, path_to_save='./csv_files', sheet_num=0):
        try:
            os.makedirs(path_to_save)
        except FileExistsError:
            pass

        path_source_file = None
        for file in self.list_files_to_read:
            if file.split(sep='/')[-1] == source_file:
                path_source_file = file

        if path_source_file:
            to_csv_file = f'{path_to_save}/{source_file}.csv'
            read_excel(path_source_file, sheet_name=sheet_num).to_csv(to_csv_file)
            print(f'"{source_file}" copied to CSV format!')
        else:
            print(f'File "{source_file}" not found!')


def ingosstrakh_pars(show_policies=False, show_data=False, save=False):
    exclude_column = (8, 9, 13, 16, 17)

    dict_to_write = {
        'Номер полиса': 0,
        'Фамилия': 1,
        'Имя': 2,
        'Отчество': 3,
        'Дата рождения': 4,
        'Пол': 5,
        'Адрес проживания': 6,
        'Наименование программы': 7,
        'Дата прикрепления': 8,
        'Дата окончания': 9,
        'Расширение': 10,
        'Ограничение': 11,
        'СНИЛС': 12,
        'Место работы': 16,
    }

    extra_cell = {
        '8 1': False,
    }

    Parser(folder_to_read='ингосстрах',
           dict_to_write=dict_to_write,
           start_line_to_read=12,
           start_column_to_read=1,
           exclude_column=exclude_column,
           extra_cell=extra_cell,
           show_policies=show_policies,
           show_data=show_data,
           save=save).pars()


def cogaz_pars(show_policies=False, show_data=False, save=False):
    dict_to_write = {
        'Фамилия': 0,
        'Имя': 1,
        'Отчество': 2,
        'Дата рождения': 3,
        'Пол': 4,
        'Адрес проживания': 5,
        'Телефон пациента': 6,
        'Номер полиса': 7,
        'Дата прикрепления': 8,
        'Дата окончания': 9,
        'Наименование программы': 10,
        'Место работы': 11,
    }

    sep_column = {
        1: None,
    }

    Parser(folder_to_read='согаз',
           dict_to_write=dict_to_write,
           start_line_to_read=20,
           start_column_to_read=1,
           exclude_column=[11],
           sep_column=sep_column,
           show_policies=show_policies,
           show_data=show_data,
           save=save).pars()


def reso_pars(show_policies=False, show_data=False, save=False):
    dict_to_write = {
        'Фамилия': 0,
        'Имя': 1,
        'Отчество': 2,
        'Дата рождения': 3,
        'Пол': 4,
        'Адрес проживания': 5,
        'Телефон пациента': 6,
        'Номер полиса': 7,
        'Дата прикрепления': 8,
        'Дата окончания': 9,
        'Наименование программы': 10,
        'Место работы': 11,
    }

    sep_column = {
        2: None,
    }

    Parser(folder_to_read='ресо',
           dict_to_write=dict_to_write,
           start_line_to_read=7,
           start_column_to_read=2,
           exclude_column=[12],
           sep_column=sep_column,
           show_policies=show_policies,
           show_data=show_data,
           save=save).pars()


def rosgosstrakh_pars(show_policies=False, show_data=False, save=False):
    dict_to_write = {
        'Фамилия': 0,
        'Имя': 1,
        'Отчество': 2,
        'Пол': 3,
        'Дата рождения': 4,
        'Адрес проживания': 5,
        'Телефон пациента': 6,
        'Номер полиса': 7,
    }

    sep_column = {
        2: None,
    }

    Parser(folder_to_read='росгострах',
           dict_to_write=dict_to_write,
           start_line_to_read=6,
           start_column_to_read=2,
           sep_column=sep_column,
           step_line=3,
           show_policies=show_policies,
           show_data=show_data,
           save=save).pars()


def alfa_pars(show_policies=False, show_data=False, save=False):
    dict_to_write = {
        'Номер полиса': 0,
        'Фамилия': 1,
        'Имя': 2,
        'Отчество': 3,
        'Дата рождения': 4,
        'Адрес проживания': 5,
        'Место работы': 6,
        'Дата прикрепления': 7,
        'Дата окончания': 8,
        'Наименование программы': 9,
        'Пол': 10,
    }

    sep_column = {
        2: None
    }

    Parser(folder_to_read='альфа',
           dict_to_write=dict_to_write,
           start_line_to_read=7,
           start_column_to_read=1,
           sep_column=sep_column,
           step_line=9,
           show_policies=show_policies,
           show_data=show_data,
           save=save).pars()


def renaissance_pars(show_policies=False, show_data=False, save=False):
    dict_to_write = {
        'Фамилия': 0,
        'Имя': 1,
        'Отчество': 2,
        'Дата рождения': 3,
        'Адрес проживания': 4,
        'Телефон пациента': 5,
        'Номер полиса': 6,
        'Дата прикрепления': 8,
        'Дата окончания': 11,
        'Наименование программы': 13,
        'Место работы': 14,
        'Пол': 15,
    }

    sep_column = {
        1: None,
    }

    extra_cell = {
        '16 2': True,
        '18 2': False,
        '14 2': False,
    }

    Parser(folder_to_read='ренессанс',
           dict_to_write=dict_to_write,
           start_line_to_read=20,
           start_column_to_read=0,
           exclude_column=[0, 3],
           sep_column=sep_column,
           extra_cell=extra_cell,
           show_policies=show_policies,
           show_data=show_data,
           save=save).pars()


def consent_pars(show_policies=False, show_data=False, save=False):
    dict_to_write = {
        'Номер полиса': 0,
        'Фамилия': 1,
        'Имя': 2,
        'Отчество': 3,
        'Дата рождения': 4,
        'Адрес проживания': 5,
        'Телефон пациента': 6,
        'Место работы': 7,
        'Дата прикрепления': 8,
        'Дата окончания': 9,
        'Наименование программы': 10,
        'Пол': 11,
    }

    sep_column = {
        3: None,
        5: '8-',
    }

    Parser(folder_to_read='согласие',
           dict_to_write=dict_to_write,
           start_line_to_read=11,
           start_column_to_read=2,
           exclude_column=[10],
           sep_column=sep_column,
           step_line=14,
           show_policies=show_policies,
           show_data=show_data,
           save=save).pars()


def alliance_pars(show_policies=False, show_data=False, save=False):
    dict_to_write = {
        'Номер полиса': 0,
        'Фамилия': 1,
        'Имя': 2,
        'Отчество': 3,
        'Дата рождения': 4,
        'Адрес проживания': 5,
        'Телефон пациента': 6,
        'Место работы': 7,
        'Дата прикрепления': 9,
        'Дата окончания': 11,
        'Наименование программы': 12,
        'Пол': 13,
    }

    sep_column = {
        3: None,
    }

    extra_cell = {
        '9 3': False,
        '11 3': True,
        '14 1': False,
    }

    Parser(folder_to_read='альянс',
           dict_to_write=dict_to_write,
           start_line_to_read=16,
           start_column_to_read=2,
           exclude_column=[5, 8],
           sep_column=sep_column,
           step_line=14,
           extra_cell=extra_cell,
           show_policies=show_policies,
           show_data=show_data,
           save=save).pars()


def parse_files(show_policies=False, show_data=False, save=False):
    ingosstrakh_pars(show_policies=show_policies, show_data=show_data, save=save)
    cogaz_pars(show_policies=show_policies, show_data=show_data, save=save)
    reso_pars(show_policies=show_policies, show_data=show_data, save=save)
    rosgosstrakh_pars(show_policies=show_policies, show_data=show_data, save=save)
    alfa_pars(show_policies=show_policies, show_data=show_data, save=save)
    renaissance_pars(show_policies=show_policies, show_data=show_data, save=save)
    consent_pars(show_policies=show_policies, show_data=show_data, save=save)
    alliance_pars(show_policies=show_policies, show_data=show_data, save=save)
    print('Pars DONE!')
    print()


def main(num_runs=1, show_policies=False, show_data=False, save=True):
    print(f'num_runs = {num_runs}')
    print()

    for i in range(num_runs):
        parse_files(show_policies=show_policies, show_data=show_data, save=save)


if __name__ == '__main__':
    main()
